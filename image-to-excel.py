import os
from azure.ai.vision.imageanalysis import ImageAnalysisClient
from azure.ai.vision.imageanalysis.models import VisualFeatures
from azure.core.credentials import AzureKeyCredential

from openpyxl import Workbook

import re

# Set the values of your computer vision endpoint and computer vision key
# as environment variables:
try:
    endpoint = "https://mikeai.cognitiveservices.azure.com/"
    key = "73b20d0d33ee42ada24eb120e0f252ac"
except KeyError:
    print("Missing environment variable 'VISION_ENDPOINT' or 'VISION_KEY'")
    print("Set them before running this sample.")
    exit()

# Create an Image Analysis client
client = ImageAnalysisClient(
    endpoint=endpoint,
    credential=AzureKeyCredential(key)
)

wb = Workbook()

ws1 = wb["Sheet"]
ws1["A1"] = "Bank"
ws1["B1"] = "Invoice Date"
ws1["C1"] = "Receipt Date"
ws1["D1"] = "Invoice No"
ws1["E1"] = "Customer Name"
ws1["F1"] = "Purpose"
ws1["G1"] = "In"
ws1["H1"] = "Out"
ws1["I1"] = "Amount"

row_start = 2

startRecord = False
skipDate = False
lastRow = False
firstRow = False
purposeList = []
search_value = '地址'
date_pattern = re.compile(r'\d{4}/\d{2}/\d{2}')
amount_pattern = re.compile(r'\d*\.\d{2}')

totalAmount = 0
amountCalculation = [0.0, 0.0]
amountCount = 0

directory = "account_data"
for filename in os.listdir(directory):
    f = os.path.join(directory, filename)
    # checking if it is a file
    if os.path.isfile(f):
        print(f)
        with open(f, "rb") as f:
            image = f.read()
        result = client.analyze(
            image_data=image,
            visual_features=[VisualFeatures.CAPTION, VisualFeatures.READ],
            gender_neutral_caption=True,  # Optional (default is False)
        ) 

        if result.read is not None:
            for line in result.read.blocks[0].lines:
                if (line.text == "往來賬戶"):
                    startRecord = True
                    firstRow = True
                    continue
                else:
                    if (startRecord == True and firstRow == True):
                        if (date_pattern.search(line.text)):
                            if (skipDate == False):
                                ws1.cell(row_start, 2).value = line.text
                            else:
                                skipDate = False
                                continue
                        elif (line.text == "承前結餘"):
                            purposeList.append(line.text)
                        elif (amount_pattern.search(line.text)):
                            totalAmount = float(line.text.replace(",", ""))
                            ws1.cell(row_start, 9).value = totalAmount
                            firstRow = False
                            continue
                    elif (startRecord == True and firstRow == False):
                        if (line.text == "月結單日期"):
                            skipDate = True
                        if (date_pattern.search(line.text)):
                            if (skipDate == False):
                                # Find the index of the first occurrence of search_value
                                index_of_search_value = -1
                                for i, item in enumerate(purposeList):
                                    if search_value in item:
                                        index_of_search_value = i
                                        break
                                # Remove elements starting from the index of the first occurrence of search_value
                                if index_of_search_value!= -1:
                                    purposeList = purposeList[:index_of_search_value]
                                ws1.cell(row_start, 6).value = " ".join(purposeList)
                                purposeList.clear()
                                row_start += 1
                                dates = line.text.split(" ")
                                if (len(dates) == 1):
                                    ws1.cell(row_start, 2).value = dates[0]
                                    lastRow = True
                                    continue
                                else: 
                                    ws1.cell(row_start, 2).value = dates[0]
                                    ws1.cell(row_start, 3).value = dates[1]
                            else:
                                skipDate = False
                                continue
                        elif (amount_pattern.search(line.text)):
                            if (lastRow == True): 
                                ws1.cell(row_start, 9).value = float(line.text.replace(",", ""))
                                # Find the index of the first occurrence of search_value
                                index_of_search_value = -1
                                for i, item in enumerate(purposeList):
                                    if search_value in item:
                                        index_of_search_value = i
                                        break
                                # Remove elements starting from the index of the first occurrence of search_value
                                if index_of_search_value!= -1:
                                    purposeList = purposeList[:index_of_search_value]
                                ws1.cell(row_start, 6).value = " ".join(purposeList)
                                purposeList.clear()
                                break
                            else: 
                                amountCalculation[amountCount] = float(line.text.replace(",", ""))
                                if (amountCount == 0):
                                    amountCount += 1
                                    continue
                                else:
                                    amountCalculation.sort()
                                    if (amountCalculation[0] + amountCalculation[1] == totalAmount):
                                        ws1.cell(row_start, 8).value = amountCalculation[0]
                                    else:
                                        ws1.cell(row_start, 7).value = amountCalculation[0]
                                    ws1.cell(row_start, 9).value = amountCalculation[1]
                                    totalAmount = amountCalculation[1]
                                    
                                    amountCount = 0
                                    continue
                        else:
                            purposeList.append(line.text)

wb.save("book_eg.xlsx")

        